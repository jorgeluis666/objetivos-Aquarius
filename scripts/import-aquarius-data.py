#!/usr/bin/env python3
"""Importa la data publicitaria mensual de Aquarius al JSON del dashboard.

Formatos soportados:

1. Tabla de resultados por campana (.csv, .xlsx, .xlsm)
   Campaña | Coste | % Δ | CTR | % Δ | Clics | % Δ | Conv | % Δ | Cos/con | % Δ

2. Serie temporal de impresiones (.csv)
   Fecha | Impresiones

Cada importacion se guarda dentro del mes correspondiente y respeta los meses
ya cargados, de modo que el filtro del dashboard acumule un mes tras otro.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "aquarius-lima-retail-2026.json"

EXPECTED_HEADERS = ["Campaña", "Coste", "% Δ", "CTR", "% Δ", "Clics", "% Δ", "Conv", "% Δ", "Cos/con", "% Δ"]
DATE_HEADERS = {"fecha", "dia", "date", "day"}
# Cada columna diaria reconocida se guarda con su nombre interno y su parser.
DAILY_METRICS = {
    "impresiones": ("impressions", "int"),
    "impressions": ("impressions", "int"),
    "coste": ("cost", "float"),
    "costo": ("cost", "float"),
    "inversion": ("cost", "float"),
    "importe gastado": ("cost", "float"),
    "gasto": ("cost", "float"),
    "resultados": ("conversions", "int"),
    "conversaciones": ("conversions", "int"),
    "conversiones": ("conversions", "int"),
    "conv": ("conversions", "int"),
    "mensajes": ("conversions", "int"),
    "clics": ("clicks", "int"),
    "clicks": ("clicks", "int"),
}
FIELDS = [
    "campaign",
    "cost",
    "costDelta",
    "ctr",
    "ctrDelta",
    "clicks",
    "clicksDelta",
    "conversions",
    "conversionsDelta",
    "costPerConversion",
    "costPerConversionDelta",
]
NUMERIC_FIELDS = set(FIELDS) - {"campaign"}

MONTH_NAMES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Setiembre", "Octubre", "Noviembre", "Diciembre",
]
MONTH_ABBR = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "set": 9, "oct": 10, "nov": 11, "dic": 12,
}


def normalize(text: object) -> str:
    value = str(text or "").strip().lower()
    replacements = str.maketrans("áéíóúüñδ", "aeiouund")
    return re.sub(r"\s+", " ", value.translate(replacements))


def parse_number(value: object) -> float | int | None:
    text = str(value or "").strip()
    if not text or text == "-":
        return None
    cleaned = re.sub(r"(s/|%)", "", text, flags=re.I).replace(",", "").strip()
    try:
        number = float(cleaned)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return int(number) if number.is_integer() else number


def parse_int_es(value: object) -> int | None:
    """Lee enteros en formato peruano: 1.234 -> 1234, 1 234 -> 1234."""
    text = str(value or "").strip()
    if not text or text == "-":
        return None
    cleaned = re.sub(r"[^\d-]", "", text)
    if not cleaned or cleaned == "-":
        return None
    return int(cleaned)


def parse_date_es(value: object) -> str | None:
    """Lee fechas tipo 'sáb, 1 ago 2026' y devuelve 2026-08-01."""
    text = normalize(value)
    match = re.search(r"(\d{1,2})\s+([a-z]{3,10})\.?\s+(\d{4})", text)
    if match:
        day, month_text, year = match.groups()
        month = MONTH_ABBR.get(month_text[:3])
        if month:
            return f"{int(year):04d}-{month:02d}-{int(day):02d}"
    match = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if match:
        year, month, day = match.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    return None


def month_label(month_id: str) -> str:
    match = re.fullmatch(r"(\d{4})-(\d{2})", month_id or "")
    if not match:
        return month_id or "Sin mes"
    year, month = match.groups()
    return f"{MONTH_NAMES[int(month) - 1]} {year}"


def month_from_filename(path: Path) -> str | None:
    match = re.search(r"(\d{4})[.\-/](\d{2})[.\-/]\d{2}", path.name)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    return None


def read_csv(path: Path) -> tuple[list[str], list[list[object]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        return [], []
    return [str(cell or "").strip() for cell in rows[0]], rows[1:]


def read_xlsx(path: Path) -> tuple[list[str], list[list[object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Para leer .xlsx instala openpyxl o exporta la fuente como CSV.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(cell or "").strip() for cell in rows[0]]
    return headers, [list(row) for row in rows[1:] if any(value not in (None, "") for value in row)]


def read_source(path: Path) -> tuple[list[str], list[list[object]]]:
    if path.suffix.lower() == ".csv":
        return read_csv(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    raise SystemExit("Formato no soportado. Usa .csv, .xlsx o .xlsm.")


def compatible(headers: list[str]) -> bool:
    if len(headers) < len(EXPECTED_HEADERS):
        return False
    return all(normalize(headers[index]) == normalize(expected) for index, expected in enumerate(EXPECTED_HEADERS))


def daily_columns(headers: list[str]) -> dict[int, tuple[str, str]]:
    """Mapea las columnas diarias reconocidas: indice -> (campo, tipo)."""
    if len(headers) < 2 or normalize(headers[0]) not in DATE_HEADERS:
        return {}
    columns = {}
    for index, header in enumerate(headers[1:], start=1):
        metric = DAILY_METRICS.get(normalize(header))
        if metric:
            columns[index] = metric
    return columns


def build_records(rows: list[list[object]]) -> list[dict[str, object]]:
    records = []
    for source in rows:
        record = {}
        for index, field in enumerate(FIELDS):
            value = source[index] if index < len(source) else None
            record[field] = parse_number(value) if field in NUMERIC_FIELDS else str(value or "").strip()
        if record.get("campaign"):
            records.append(record)
    return records


def build_daily_rows(rows: list[list[object]], columns: dict[int, tuple[str, str]]) -> list[dict[str, object]]:
    daily = []
    for source in rows:
        if not source:
            continue
        date = parse_date_es(source[0])
        if not date:
            continue
        entry = {"date": date}
        for index, (field, kind) in columns.items():
            value = source[index] if index < len(source) else None
            parsed = parse_int_es(value) if kind == "int" else parse_number(value)
            if parsed is not None:
                entry[field] = parsed
        if len(entry) > 1:
            daily.append(entry)
    daily.sort(key=lambda item: item["date"])
    return daily


def merge_daily(month: dict[str, object], rows: list[dict[str, object]], source_name: str) -> None:
    """Agrega o actualiza los dias del mes sin borrar metricas ya cargadas."""
    daily = month.get("daily")
    if not isinstance(daily, dict):
        daily = {"sourceFiles": [], "rows": []}
    existing = {item["date"]: item for item in daily.get("rows", [])}
    for row in rows:
        existing.setdefault(row["date"], {"date": row["date"]}).update(row)
    daily["rows"] = sorted(existing.values(), key=lambda item: item["date"])
    sources = [name for name in daily.get("sourceFiles", []) if name != source_name]
    daily["sourceFiles"] = sources + [source_name]
    month["daily"] = daily
    month.pop("impressions", None)


def load_document(output: Path) -> dict[str, object]:
    """Lee el JSON existente y lo migra al esquema por meses si hace falta."""
    base = {
        "brand": "Aquarius",
        "dashboard": "Gasto Publicitario",
        "moduleSubtitle": "Branding y ventas",
        "schemaVersion": 2,
        "status": "ok",
        "defaultMonth": None,
        "months": [],
    }
    if not output.exists():
        return base
    try:
        current = json.loads(output.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return base
    for key in ("brand", "dashboard", "moduleSubtitle"):
        if current.get(key):
            base[key] = current[key]
    if isinstance(current.get("months"), list):
        base["months"] = [migrate_month(month) for month in current["months"]]
        base["defaultMonth"] = current.get("defaultMonth")
        return base
    if isinstance(current.get("records"), list) and current["records"]:
        legacy_id = current.get("month") or "historico"
        base["months"] = [{
            "id": legacy_id,
            "label": month_label(legacy_id) if re.fullmatch(r"\d{4}-\d{2}", legacy_id) else "Historico",
            "sourceFile": current.get("sourceFile"),
            "receivedHeaders": current.get("receivedHeaders", []),
            "records": current["records"],
        }]
        base["defaultMonth"] = legacy_id
    return base


def migrate_month(month: dict[str, object]) -> dict[str, object]:
    """Convierte el bloque `impressions` antiguo en la serie diaria unificada."""
    legacy = month.get("impressions")
    if isinstance(legacy, dict) and isinstance(legacy.get("daily"), list):
        rows = [
            {"date": item["date"], "impressions": item["impressions"]}
            for item in legacy["daily"]
            if item.get("date") is not None and item.get("impressions") is not None
        ]
        merge_daily(month, rows, legacy.get("sourceFile") or "serie-impresiones")
    return month


def upsert_month(document: dict[str, object], month_id: str) -> dict[str, object]:
    for month in document["months"]:
        if month.get("id") == month_id:
            return month
    month = {"id": month_id, "label": month_label(month_id), "records": []}
    document["months"].append(month)
    return month


def mirror_legacy_month(document: dict[str, object]) -> None:
    """Copia el mes por defecto a la raiz del JSON.

    Sirve de compatibilidad: un dashboard con el JS anterior en cache sigue
    leyendo `records` y no se rompe mientras GitHub Pages refresca sus assets.
    """
    months = document.get("months") or []
    default = next((month for month in months if month.get("id") == document.get("defaultMonth")), None)
    if not default:
        document.pop("records", None)
        return
    document["month"] = default.get("id")
    document["sourceFile"] = default.get("sourceFile")
    document["receivedHeaders"] = default.get("receivedHeaders", [])
    document["records"] = default.get("records", [])


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa data publicitaria mensual de Aquarius.")
    parser.add_argument("source", type=Path, help="Archivo fuente .csv, .xlsx o .xlsm")
    parser.add_argument("--month", help="Mes destino en formato AAAA-MM, por ejemplo 2026-09")
    parser.add_argument("--label", help="Etiqueta visible del mes, por ejemplo 'Setiembre 2026'")
    parser.add_argument("--output", type=Path, default=OUTPUT, help="Ruta JSON de salida")
    args = parser.parse_args()

    headers, rows = read_source(args.source)
    document = load_document(args.output)

    columns = daily_columns(headers)
    if not compatible(headers) and not columns:
        print("[aquarius-import] el archivo no coincide con los formatos esperados.")
        print(f"  encabezados recibidos: {headers}")
        print(f"  tabla de campanas: {EXPECTED_HEADERS}")
        print("  serie diaria: primera columna Fecha y alguna de " + ", ".join(sorted({name for name, _ in DAILY_METRICS.values()})))
        return 1

    month_id = args.month or month_from_filename(args.source)
    if columns and not compatible(headers):
        daily = build_daily_rows(rows, columns)
        if not daily:
            print("[aquarius-import] la serie diaria no tiene filas legibles.")
            return 1
        month_id = month_id or daily[0]["date"][:7]
        month = upsert_month(document, month_id)
        merge_daily(month, daily, args.source.name)
        metrics = sorted({field for field, _ in columns.values()})
        detail = f"{len(daily)} dias con {', '.join(metrics)}"
    else:
        if not month_id:
            print("[aquarius-import] indica el mes destino con --month AAAA-MM.")
            return 1
        month = upsert_month(document, month_id)
        month["sourceFile"] = args.source.name
        month["receivedHeaders"] = headers
        month["records"] = build_records(rows)
        detail = f"{len(month['records'])} campanas"

    if args.label:
        month["label"] = args.label
    elif not month.get("label"):
        month["label"] = month_label(month_id)

    document["months"].sort(key=lambda item: str(item.get("id") or ""))
    document["defaultMonth"] = document["months"][-1]["id"] if document["months"] else None
    document["status"] = "ok"
    mirror_legacy_month(document)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[aquarius-import] {month['label']}: {detail} -> {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
